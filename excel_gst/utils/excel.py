from openpyxl.styles import Font, Alignment
from datetime import datetime
import config

# Generic check method
def check(worksheet, row: int, check_map: dict):
    """
    A generic check method used to validate whether to count a particular row or not in selection based on a check map

    Parameters:

    worksheet : Worksheet : Worksheet to which row belongs to

    row : int : Row to be checked

    check_map : dict : A dictionary mapping of column numbers to values. Example: {3: "Debit note"} means 3rd column of this particular row needs to be equal to the string 'Debit note'. When multiple checks are passed, all of them need to pass for it to be True.

    Returns: Boolean
    """
    boolean_list = []
    for column, value in check_map.items():
        cell = worksheet.cell(row, column)
        boolean_list.append(cell.value == value)
    return all(boolean_list)


def selectRows(workbook, sheet_name: str, check_map: dict):
    """
    Select rows of a particular worksheet based on the checks given

    Parameters:

    sheet_name : str : Name of the worksheet. Example: B2B

    check_map : dict : A dictionary mapping of column numbers to values. Example: {3: "Debit note"} means 3rd column of this particular row needs to be equal to the string 'Debit note'. When multiple checks are passed, all of them need to pass for it to be True.

    Returns: list : First element is a list of row numbers that are selected and second element is the worksheet object
    """
    requiredRows = []

    worksheet = workbook[sheet_name]
    maxRows = worksheet.max_row

    for row_number in range(1, maxRows + 1):
        if check(worksheet, row_number, check_map):
            requiredRows.append(row_number)

    return [requiredRows, worksheet]


def getSumOfRowValues(column: int, requiredRows: list, worksheet):
    """
    Returns sum of the selected rows

    Parameters:

    column : int : The column of the rows to be added

    requiredRows : list : List of row numbers that needs to be counted.

    worksheet : Worksheet : The worksheet object where the column and rows belong to.

    Returns: int : The sum of values in selected rows
    """
    total = 0
    for row in requiredRows:
        cell = worksheet.cell(row, column)
        total += cell.value
    return total


def resolve_headings(worksheet):
    for heading, start_col in config.WORKSHEET_PRIMARY_HEADINGS.items():
        worksheet.merge_cells(
            start_row=1,
            end_row=1,
            start_column=start_col,
            end_column=start_col + config.COLUMNS_LENGTH - 1,
        )
        top_left_cell = worksheet.cell(1, start_col)
        top_left_cell.value = heading
        top_left_cell.font = Font(b=True, sz=20)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
    for column, secondary_heading in config.WORKSHEET_SECONDARY_HEADINGS.items():
        sec_cell = worksheet.cell(row=2, column=column, value=secondary_heading)
        sec_cell.font = Font(b=True, sz=15)
        sec_cell.alignment = Alignment(horizontal="center", vertical="center")
    month_heading = worksheet.cell(2, 1, "Month")
    month_heading.alignment = Alignment(horizontal="center", vertical="center")
    month_heading.font = Font(b=True, sz=15)


def insertmonth(workbook, worksheet, row, string_month=None):
    if not string_month:
        readme = workbook["Read me"]
        month = readme.cell(2, 5)
        string_month = str(month.value)
    month_string, year_string = string_month[:2], string_month[2:]
    datetime_obj = datetime(year=int(year_string), month=int(month_string), day=1)
    month_cell = worksheet.cell(row, 1, datetime_obj)
    month_cell.number_format = "MMM-YYYY"


def calculator(
    workbook,
    sheet_name: str,
    column_map: dict,
    check_map: dict,
    worksheet,
    start_row: int,
    start_col: int,
):
    """
    Main calculator function

    Parameters:

    sheet_name : str : Name of the sheet

    column_map : dict : Map of column number to column readable name. All the suitable rows of these particular columns will be added.

    check_map : dict : A dictionary mapping of column numbers to values. Example: {3: "Debit note"} means 3rd column of this particular row needs to be equal to the string 'Debit note'. When multiple checks are passed, all of them need to pass for it to be True.

    Returns: void : Prints the calculations
    """
    rows, original_worksheet = selectRows(workbook, sheet_name, check_map)
    column = start_col
    for (key, value) in column_map.items():
        answer = getSumOfRowValues(key, rows, original_worksheet)
        result = round(answer, 2)
        value_cell = worksheet.cell(start_row, column, value=result)
        value_cell.font = Font(sz=12)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        column += 1
